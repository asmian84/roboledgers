#!/usr/bin/env python3
"""
Extract description → COA account mappings from all XLSX files in /Users/asmian/XLSX/
Scans every sheet in every file for columns that look like "description" and "account"
Outputs a consolidated JSON mapping file for RoboLedger's categorization engine.
"""

import os
import json
import re
import sys
from collections import defaultdict, Counter
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

XLSX_DIR = "/Users/asmian/XLSX"
OUTPUT_FILE = "/Users/asmian/RoboLedger - working copy 2/.claude/worktrees/hungry-villani/src/data/firm_categorizations.json"

# Column name patterns to match
DESC_PATTERNS = re.compile(r'desc|narrat|detail|particular|memo|payee|vendor|transaction|note', re.I)
ACCT_PATTERNS = re.compile(r'account|acct|acc\b|coa|code|gl|ledger|category|cat\b|map', re.I)
# Exclude columns that are clearly not what we want
EXCLUDE_PATTERNS = re.compile(r'balance|running|total|date|amount|debit|credit|ref|cheque|check', re.I)

def find_columns(headers):
    """Find description and account column indices from headers."""
    desc_col = None
    acct_col = None

    for i, h in enumerate(headers):
        if h is None:
            continue
        h_str = str(h).strip()
        if not h_str:
            continue

        # Skip excluded patterns
        if EXCLUDE_PATTERNS.search(h_str) and not ACCT_PATTERNS.search(h_str):
            continue

        if DESC_PATTERNS.search(h_str) and desc_col is None:
            desc_col = i
        elif ACCT_PATTERNS.search(h_str) and acct_col is None:
            acct_col = i

    return desc_col, acct_col

def clean_description(desc):
    """Normalize transaction description for pattern matching."""
    if not desc:
        return None
    desc = str(desc).strip()
    if len(desc) < 3 or len(desc) > 200:
        return None
    # Remove excessive whitespace
    desc = re.sub(r'\s+', ' ', desc)
    return desc

def clean_account(acct):
    """Extract numeric COA account code."""
    if acct is None:
        return None
    acct = str(acct).strip()

    # Extract numeric portion (4-digit COA codes)
    match = re.search(r'\b(\d{4})\b', acct)
    if match:
        code = match.group(1)
        # Valid COA range: 1000-9999
        num = int(code)
        if 1000 <= num <= 9999:
            return code

    # Try pure numeric
    if acct.isdigit() and len(acct) == 4:
        num = int(acct)
        if 1000 <= num <= 9999:
            return acct

    return None

def extract_from_file(filepath):
    """Extract description → account mappings from a single XLSX file."""
    mappings = []
    filename = os.path.basename(filepath)

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return mappings

    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            if len(rows) < 2:
                continue

            # Try first few rows as potential headers
            for header_row_idx in range(min(5, len(rows))):
                headers = rows[header_row_idx]
                if headers is None:
                    continue

                desc_col, acct_col = find_columns(headers)

                if desc_col is not None and acct_col is not None:
                    # Found both columns - extract data rows
                    for row in rows[header_row_idx + 1:]:
                        if row is None or len(row) <= max(desc_col, acct_col):
                            continue

                        desc = clean_description(row[desc_col])
                        acct = clean_account(row[acct_col])

                        if desc and acct:
                            mappings.append({
                                'description': desc,
                                'account': acct,
                                'source': filename,
                                'sheet': sheet_name
                            })
                    break  # Found valid headers, don't try other rows

        except Exception as e:
            continue

    try:
        wb.close()
    except:
        pass

    return mappings

def generate_vendor_rules(mappings):
    """Analyze mappings to generate vendor → account rules with confidence."""

    # Group by normalized description → account votes
    desc_to_accounts = defaultdict(lambda: Counter())
    desc_examples = defaultdict(list)

    for m in mappings:
        desc_upper = m['description'].upper()
        # Normalize: remove dates, amounts, reference numbers
        normalized = re.sub(r'\d{4}[-/]\d{2}[-/]\d{2}', '', desc_upper)
        normalized = re.sub(r'\$[\d,]+\.?\d*', '', normalized)
        normalized = re.sub(r'#\d+', '', normalized)
        normalized = re.sub(r'\b\d{6,}\b', '', normalized)  # Long numbers
        normalized = re.sub(r'\s+', ' ', normalized).strip()

        if len(normalized) < 3:
            continue

        desc_to_accounts[normalized][m['account']] += 1
        if len(desc_examples[normalized]) < 3:
            desc_examples[normalized].append(m['description'])

    # Generate rules from consistent mappings
    rules = []
    for desc, account_votes in desc_to_accounts.items():
        total_votes = sum(account_votes.values())
        if total_votes < 1:
            continue

        top_account, top_count = account_votes.most_common(1)[0]
        confidence = top_count / total_votes

        rules.append({
            'pattern': desc,
            'account': top_account,
            'confidence': round(confidence, 3),
            'occurrences': total_votes,
            'examples': desc_examples[desc][:3],
            'alternatives': dict(account_votes) if len(account_votes) > 1 else None
        })

    # Sort by occurrences (most common first)
    rules.sort(key=lambda r: (-r['occurrences'], -r['confidence']))

    return rules

def generate_keyword_rules(rules):
    """Extract keyword → account patterns from high-confidence rules."""

    # Common vendor/keyword patterns
    keyword_accounts = defaultdict(lambda: Counter())

    for rule in rules:
        if rule['confidence'] < 0.7 or rule['occurrences'] < 2:
            continue

        # Extract significant words (3+ chars, not common words)
        words = re.findall(r'[A-Z]{3,}', rule['pattern'])
        stop_words = {'THE', 'AND', 'FOR', 'FROM', 'WITH', 'PAYMENT', 'PURCHASE',
                      'TRANSACTION', 'TRANSFER', 'DEPOSIT', 'WITHDRAWAL', 'DEBIT',
                      'CREDIT', 'ONLINE', 'BANKING', 'INTERNET', 'MOBILE', 'POS',
                      'PRE', 'AUTH', 'TXN', 'REF', 'INV', 'PAY', 'REC', 'RECEIVED',
                      'SENT', 'INTERAC', 'VISA', 'MASTERCARD', 'AMEX', 'CANADA'}

        keywords = [w for w in words if w not in stop_words and len(w) >= 3]

        for kw in keywords:
            keyword_accounts[kw][rule['account']] += rule['occurrences']

    # Generate keyword rules
    keyword_rules = []
    for kw, account_votes in keyword_accounts.items():
        total = sum(account_votes.values())
        if total < 3:
            continue
        top_account, top_count = account_votes.most_common(1)[0]
        confidence = top_count / total

        if confidence >= 0.6:
            keyword_rules.append({
                'keyword': kw,
                'account': top_account,
                'confidence': round(confidence, 3),
                'total_occurrences': total
            })

    keyword_rules.sort(key=lambda r: (-r['total_occurrences'], -r['confidence']))
    return keyword_rules

def main():
    print(f"Scanning {XLSX_DIR} for categorized transaction data...")

    xlsx_files = sorted(Path(XLSX_DIR).glob("*.xlsx"))
    print(f"Found {len(xlsx_files)} Excel files")

    all_mappings = []
    files_with_data = 0
    files_scanned = 0

    for filepath in xlsx_files:
        files_scanned += 1
        if files_scanned % 50 == 0:
            print(f"  Scanned {files_scanned}/{len(xlsx_files)} files... ({len(all_mappings)} mappings so far)")

        mappings = extract_from_file(str(filepath))
        if mappings:
            all_mappings.extend(mappings)
            files_with_data += 1

    print(f"\nResults:")
    print(f"  Files scanned: {files_scanned}")
    print(f"  Files with data: {files_with_data}")
    print(f"  Total mappings: {len(all_mappings)}")

    if not all_mappings:
        print("No mappings found! Check column names in the Excel files.")
        # Show sample headers from first few files
        print("\nSample headers from first 5 files:")
        for filepath in xlsx_files[:5]:
            try:
                wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
                for sheet in wb.sheetnames[:1]:
                    ws = wb[sheet]
                    for row in ws.iter_rows(max_row=3, values_only=True):
                        print(f"  {filepath.name} [{sheet}]: {[str(c)[:30] if c else '' for c in row]}")
                wb.close()
            except:
                pass
        return

    # Account distribution
    account_counts = Counter(m['account'] for m in all_mappings)
    print(f"\n  Unique accounts used: {len(account_counts)}")
    print(f"\n  Top 20 accounts by usage:")
    for acct, count in account_counts.most_common(20):
        print(f"    {acct}: {count} transactions")

    # Generate rules
    print("\nGenerating vendor rules...")
    vendor_rules = generate_vendor_rules(all_mappings)
    print(f"  Generated {len(vendor_rules)} vendor rules")

    high_conf = [r for r in vendor_rules if r['confidence'] >= 0.9 and r['occurrences'] >= 3]
    print(f"  High confidence (≥90%, 3+ occurrences): {high_conf.__len__()} rules")

    print("\nGenerating keyword rules...")
    keyword_rules = generate_keyword_rules(vendor_rules)
    print(f"  Generated {len(keyword_rules)} keyword rules")

    # Source file distribution
    source_counts = Counter(m['source'] for m in all_mappings)
    print(f"\n  Data from {len(source_counts)} unique source files")
    print(f"  Top 10 files by mapping count:")
    for src, count in source_counts.most_common(10):
        print(f"    {src}: {count} mappings")

    # Build output
    output = {
        'metadata': {
            'total_mappings': len(all_mappings),
            'files_with_data': files_with_data,
            'files_scanned': files_scanned,
            'unique_accounts': len(account_counts),
            'generated_at': __import__('datetime').datetime.now().isoformat()
        },
        'account_distribution': dict(account_counts.most_common()),
        'vendor_rules': vendor_rules,  # ALL vendor rules
        'keyword_rules': keyword_rules[:200],  # Top 200 keyword rules
        'high_confidence_rules': high_conf,
        'raw_mappings_sample': all_mappings[:100]  # Sample for debugging
    }

    # Save
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with open(OUTPUT_FILE, 'w') as f:
        json.dump(output, f, indent=2)

    print(f"\nOutput saved to: {OUTPUT_FILE}")
    print(f"File size: {os.path.getsize(OUTPUT_FILE) / 1024:.1f} KB")

    # Print some example high-confidence rules
    if high_conf:
        print(f"\n{'='*60}")
        print(f"TOP 30 HIGH-CONFIDENCE RULES (≥90%, 3+ occurrences):")
        print(f"{'='*60}")
        for r in high_conf[:30]:
            print(f"  [{r['account']}] {r['pattern'][:60]:<60} conf={r['confidence']:.0%} n={r['occurrences']}")

if __name__ == '__main__':
    main()
